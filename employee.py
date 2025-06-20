from flask import Flask, jsonify, render_template, request, redirect, url_for, flash, session
from datetime import datetime, date
from openpyxl import Workbook
from psycopg2.extras import RealDictCursor
from openpyxl.styles import Font
import mysql.connector
import psycopg2
import pandas as pd
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

def get_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASS"),
        port=5432
    )

# ==================== ดึงรายชื่อแผนก ====================
def get_departments():
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT DISTINCT name FROM departments ORDER BY name")
    departments = [row["name"] for row in cursor.fetchall()]
    conn.close()
    return departments

    old_dept_name = get_department_name(old_dept_id)
    new_dept_name = get_department_name(new_dept_id)

    flash(f"เปลี่ยนแผนกจาก {old_dept_name} เป็น {new_dept_name}", "success")

# ==================== ฟังก์ชันจัดการวันที่ ====================
def parse_thai_date(date_str):
    """
    แปลงวันที่รูปแบบ 'dd/mm/yyyy' ปีพุทธศักราช เป็น datetime.date ปีคริสต์ศักราช
    เช่น '01/01/2530' -> datetime.date(1987, 1, 1)
    """
    try:
        if not date_str or '/' not in date_str:
            return None
        day, month, year = map(int, date_str.strip().split("/"))
        if year > 2400:
            year -= 543
        return date(year, month, day)
    except Exception as e:
        print(f"Error parsing date '{date_str}': {e}")
        return None

def format_thai_date(date_obj):
    """
    แปลง datetime.date หรือ datetime.datetime เป็นวันที่รูปแบบ 'dd/mm/yyyy' ปีพุทธศักราช
    ถ้า input ไม่ถูกต้องหรือ None คืนค่าว่าง ""
    """
    try:
        if not date_obj or str(date_obj).strip().lower() in ['none', '', 'null']:
            return ""
        if isinstance(date_obj, (datetime, date)):
            d = date_obj
        else:
            d = datetime.strptime(str(date_obj), "%Y-%m-%d")
        return d.strftime(f"%d/%m/{d.year + 543}")
    except Exception:
        return ""

def prepare_employee_data(employee_list):
    """
    แปลงและเตรียมข้อมูล employee_list ที่เป็น dict
    แปลงวันที่เป็นปีพุทธศักราชแบบ string และคำนวณอายุงาน
    """
    for emp in employee_list:
        # แปลงข้อมูลวันที่ที่เป็น string เป็น datetime.date ก่อน (ถ้ายังไม่ได้แปลง)
        if isinstance(emp.get("start_date"), str):
            emp["start_date"] = parse_thai_date(emp["start_date"])
        if isinstance(emp.get("exit_date"), str):
            emp["exit_date"] = parse_thai_date(emp["exit_date"])
        if isinstance(emp.get("birth"), str):
            emp["birth"] = parse_thai_date(emp["birth"])

        start_date = emp.get("start_date")
        exit_date = emp.get("exit_date")
        birth_date = emp.get("birth")
        print("start_date =", start_date)
        print("exit_date =", exit_date)

        # แปลงวันเริ่มงานเป็นวันที่ไทย
        if isinstance(start_date, (datetime, date)):
            emp["start_date_fmt"] = format_thai_date(start_date)
            emp["start_day"] = start_date.day
            emp["start_month"] = start_date.month
            emp["start_year"] = start_date.year + 543
        else:
            emp["start_date_fmt"] = ""
            emp["start_day"] = ""
            emp["start_month"] = ""
            emp["start_year"] = ""

        # วันเกิด (แปลงเป็น string ไทย)
        if isinstance(birth_date, (datetime, date)):
            emp["birth"] = format_thai_date(birth_date)
        else:
            emp["birth"] = ""

        # วันออก (แปลงเป็น string ไทย)
        if isinstance(exit_date, (datetime, date)):
            emp["exit_date_fmt"] = format_thai_date(exit_date)
        else:
            emp["exit_date_fmt"] = ""

    return employee_list

@app.route("/test_db")
def test_db():
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT version();")
        version = cur.fetchone()
        return f"Database connected successfully: {version}"
    except Exception as e:
        return f"Connection failed: {e}"

# ==================== หน้าแรก ====================
@app.route('/')
def index():
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # ดึงค่าค้นหาจาก URL parameters
    name = request.args.get('name', '').strip()
    gender = request.args.get('gender', '')
    dept_name = request.args.get('department', '')
    birth_from = parse_thai_date(request.args.get('birth_from'))
    birth_to = parse_thai_date(request.args.get('birth_to'))

    # สร้าง query และ parameter แบบ dynamic
    query = """
        SELECT e.*, d.name AS department_name FROM employee e
        LEFT JOIN departments d ON e.department_id = d.id
        WHERE 1=1
    """
    params = []

    if name:
        query += " AND (e.name LIKE %s OR e.sur_name LIKE %s)"
        params += [f"%{name}%", f"%{name}%"]
    if gender:
        query += " AND e.gender = %s"
        params.append(gender)
    if dept_name:
        query += " AND d.name = %s"
        params.append(dept_name)
    if birth_from and birth_to:
        query += " AND e.birth BETWEEN %s AND %s"
        params += [birth_from, birth_to]

    query += " ORDER BY id ASC"
    cursor.execute(query, params)
    employee = cursor.fetchall()

    for emp in employee:
        emp["birth"] = format_thai_date(emp["birth"])
        emp["start_date"] = format_thai_date(emp["start_date"])
        # เพิ่มการแยกวัน เดือน ปี เพื่อให้ใช้ใน template ได้
        try:
            date_parts = emp["start_date"].split("/")
            emp["start_day"] = int(date_parts[0])
            emp["start_month"] = int(date_parts[1])
            emp["start_year"] = int(date_parts[2])  # พ.ศ. แล้ว
        except Exception:
            emp["start_day"] = emp["start_month"] = emp["start_year"] = None

    cursor.execute("SELECT name FROM departments ORDER BY name")
    departments = [row['name'] for row in cursor.fetchall()]
    conn.close()

    return render_template('index.html', employee=employee, departments=departments, search_params={
        "name": name,
        "gender": gender,
        "department": dept_name,
        "birth_from": request.args.get('birth_from', ''),
        "birth_to": request.args.get('birth_to', '')
    })

# ==================== เพิ่มข้อมูลพนักงาน ====================
@app.route('/add', methods=['POST'])
def add():
    birth = parse_thai_date(request.form['birth'])
    start_date = parse_thai_date(request.form['start_date'])

    if not birth or not start_date:
        flash("รูปแบบวันเกิดหรือวันเริ่มงานไม่ถูกต้อง", "danger")
        return redirect(url_for('index'))

    try:
        # รับค่าจาก select และตรวจสอบว่าถ้าเลือก "อื่นๆ" ให้ใช้ค่าจาก input
        department = request.form['department']
        if department == 'อื่นๆ':
            department = request.form.get('other_department', '').strip()

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO employee (
                id, pre, name, sur_name, gender, type,
                department_id, id_card, nationality,
                senior, birth, phone, email, address, 
                start_date, bank_account, main_skill, sub_skill1, sub_skill2
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            request.form['id'],
            request.form['pre'],
            request.form['name'],
            request.form['sur_name'],
            request.form['gender'],
            request.form['type'],
            department,  # ใช้ค่าที่ปรับแล้ว
            request.form['id_card'],
            request.form['nationality'],
            request.form['senior'],
            birth,
            request.form['phone'],
            request.form['email'],
            request.form['address'],
            start_date,
            request.form['bank_account'],
            request.form['main_skill'],
            request.form['sub_skill1'],
            request.form['sub_skill2']
        ))
        conn.commit()
        flash('เพิ่มข้อมูลเรียบร้อยแล้ว', 'success')

    except mysql.connector.IntegrityError:
        flash('รหัสนี้มีอยู่แล้ว', 'danger')
    finally:
        conn.close()

    department = request.form['department']
    if department == 'อื่นๆ':
        new_dept = request.form.get('other_department', '').strip()
        category = request.form.get('department_category', 'ทั่วไป').strip()
        if new_dept:
            # เพิ่มแผนกใหม่ในฐานข้อมูล
            conn = get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("INSERT IGNORE INTO departments (name, category) VALUES (%s, %s)", (new_dept, category))
                conn.commit()
                department = new_dept
            except Exception as e:
                flash("เพิ่มแผนกใหม่ไม่สำเร็จ: " + str(e), "danger")
            finally:
                conn.close()

    return redirect(url_for('index'))

# ==================== API เพิ่ม/ลบ/แผนก ====================
@app.route('/api/add_department', methods=['POST'])
def add_department_api():
    dept = request.form.get('department', '').strip()
    if not dept:
        return {"success": False, "error": "ชื่อแผนกว่าง"}

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("INSERT IGNORE INTO departments (name) VALUES (%s)", (dept,))
    conn.commit()
    conn.close()
    return {"success": True}

@app.route('/api/add_employee', methods=['POST'])
def api_add_employee():
    data = request.form
    birth = parse_thai_date(data['birth'])
    start_date = parse_thai_date(data['start_date'])

    if not birth or not start_date:
        return {'success': False, 'error': 'วันเกิดหรือวันเริ่มงานไม่ถูกต้อง'}

    dept_name = data['department']
    if dept_name == 'อื่นๆ':
        dept_name = data.get('other_department', '').strip()

    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT id FROM departments WHERE name = %s", (dept_name,))
    dept_row = cursor.fetchone()

    if not dept_row:
        cursor.execute("INSERT INTO departments (name) VALUES (%s)", (dept_name,))
        conn.commit()
        cursor.execute("SELECT id FROM departments WHERE name = %s", (dept_name,))
        dept_row = cursor.fetchone()

    department_id = dept_row['id']

    try:
        cursor.execute('''
            INSERT INTO employee (
                id, pre, name, sur_name, gender, type,
                department_id, id_card, nationality, senior,
                birth, phone, email, address, start_date, bank_account,
                main_skill, sub_skill1, sub_skill2
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            data['id'], data['pre'], data['name'], data['sur_name'],
            data['gender'], data['type'], department_id,
            data['id_card'], data['nationality'], data['senior'],
            birth, data['phone'], data['email'], data['address'],
            start_date, data['bank_account'], data['main_skill'],
            data['sub_skill1'], data['sub_skill2']
        ))
        conn.commit()
        return {
            'success': True,
            'employee': {
                'id': data['id'], 'pre': data['pre'], 'name': data['name'],
                'sur_name': data['sur_name'], 'gender': data['gender'],
                'type': data['type'], 'department': dept_name,
                'id_card': data['id_card'], 'nationality': data['nationality'],
                'senior': data['senior'], 'birth': data['birth'],
                'phone': data['phone'], 'email': data['email'],
                'address': data['address'], 'start_date': data['start_date'],
                'bank_account': data['bank_account'], 'main_skill': data['main_skill'],
                'sub_skill1': data['sub_skill1'], 'sub_skill2': data['sub_skill2']
            }
        }
    except mysql.connector.IntegrityError:
        return {'success': False, 'error': 'รหัสพนักงานซ้ำ'}
    finally:
        conn.close()


@app.route('/api/delete_employee', methods=['POST'])
def delete_employee_api():  # ✅ ชื่อไม่ซ้ำกับเดิม
    emp_id = request.json.get('id')

    if not emp_id:
        return jsonify(success=False, error="ไม่พบรหัสพนักงาน"), 400

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM employee WHERE id = %s", (emp_id,))
    conn.commit()
    conn.close()

    return jsonify(success=True)

# ==================== อัปเดตข้อมูลพนักงาน ====================
@app.route('/update/<int:emp_id>', methods=['GET', 'POST'])
def update(emp_id):
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # ดึงข้อมูลพนักงานปัจจุบัน
    cursor.execute("""
        SELECT e.*, d.name AS department_name
        FROM employee e
        LEFT JOIN departments d ON e.department_id = d.id
        WHERE e.id = %s
    """, (emp_id,))
    employee = cursor.fetchone()

    # ดึงข้อมูลแผนกทั้งหมด พร้อมสร้าง dict สำหรับแปลง id -> ชื่อ
    cursor.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cursor.fetchall()
    dept_dict = {str(d["id"]): d["name"] for d in departments}

    if not employee:
        flash("ไม่พบพนักงาน", "danger")
        conn.close()
        return redirect(url_for('index'))

    # กำหนดชื่อฟิลด์ภาษาไทยสำหรับใช้แสดงผล
    field_labels = {
        "pre": "คำนำหน้า",
        "name": "ชื่อ",
        "sur_name": "นามสกุล",
        "gender": "เพศ",
        "type": "ประเภท",
        "department_id": "แผนก",
        "id_card": "เลขบัตรประชาชน",
        "nationality": "สัญชาติ",
        "senior": "วุฒิ",
        "birth": "วันเกิด",
        "phone": "เบอร์โทร",
        "email": "อีเมล",
        "address": "ที่อยู่",
        "start_date": "วันเริ่มงาน",
        "bank_account": "เลขบัญชี",
        "main_skill": "ความสามารถหลัก",
        "sub_skill1": "ความสามารถอื่น 1",
        "sub_skill2": "ความสามารถอื่น 2",
        "exit_date": "วันออก",
        "exit_reason": "หมายเหตุ"
    }

    if request.method == 'POST':
        # แปลงวันที่จากฟอร์ม
        new_birth = parse_thai_date(request.form['birth'])
        new_start = parse_thai_date(request.form['start_date'])
        exit_date_raw = request.form.get('exit_date', '').strip()

        if not new_birth or not new_start:
            flash("รูปแบบวันเกิดหรือวันเริ่มงานไม่ถูกต้อง", "danger")
            conn.close()
            return redirect(url_for('update', emp_id=emp_id))

        # แปลง exit_date เฉพาะกรณีกรอกมา
        if exit_date_raw:
            new_exit = parse_thai_date(exit_date_raw)
        else:
            new_exit = None  # จะไม่ใส่ใน SQL

        # บันทึกข้อมูลเก่าไว้ใน employee_history
        cursor.execute("""
            INSERT INTO employee_history (
                employee_id, pre, name, sur_name, gender, type, department_id, id_card, nationality,
                senior, birth, phone, email, address, start_date, bank_account, main_skill, sub_skill1, sub_skill2, exit_date, exit_reason
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            emp_id,
            employee.get('pre'), employee.get('name'), employee.get('sur_name'), employee.get('gender'),
            employee.get('type'), employee.get('department_id'), employee.get('id_card'), employee.get('nationality'),
            employee.get('senior'), employee.get('birth'), employee.get('phone'), employee.get('email'),
            employee.get('address'), employee.get('start_date'), employee.get('bank_account'),
            employee.get('main_skill'), employee.get('sub_skill1'), employee.get('sub_skill2'),
            employee.get('exit_date'), employee.get('exit_reason')
        ))

        # เก็บข้อมูลใหม่
        new_data = {
            "pre": request.form['pre'],
            "name": request.form['name'],
            "sur_name": request.form['sur_name'],
            "gender": request.form['gender'],
            "type": request.form['type'],
            "department_id": request.form.get('department_id') or None,
            "id_card": request.form['id_card'],
            "nationality": request.form['nationality'],
            "senior": request.form['senior'],
            "birth": new_birth,
            "phone": request.form['phone'],
            "email": request.form['email'],
            "address": request.form['address'],
            "start_date": new_start,
            "bank_account": request.form['bank_account'],
            "main_skill": request.form.get('main_skill'),
            "sub_skill1": request.form.get('sub_skill1'),
            "sub_skill2": request.form.get('sub_skill2'),
            "exit_reason": request.form.get("exit_reason")
        }

        # เพิ่ม exit_date หากกรอกมา
        if new_exit:
            new_data["exit_date"] = new_exit

        # ตรวจสอบการเปลี่ยนแปลง
        changes = []
        for key, new_value in new_data.items():
            old_value = employee.get(key)
            if key in ['birth', 'start_date', 'exit_date']:
                old_str = format_thai_date(old_value) if old_value else ""
                new_str = format_thai_date(new_value) if new_value else ""
            elif key == "department_id":
                old_str = dept_dict.get(str(old_value), "ไม่ระบุ")
                new_str = dept_dict.get(str(new_value), "ไม่ระบุ")
            else:
                old_str = str(old_value or "")
                new_str = str(new_value or "")
            if old_str != new_str:
                changes.append(f"{field_labels.get(key, key)}: '{old_str}' → '{new_str}'")

        # อัปเดตแบบไดนามิก
        update_fields = list(new_data.keys())
        update_clause = ", ".join([f"{field} = %s" for field in update_fields])
        values = [new_data[field] for field in update_fields]
        values.append(emp_id)

        cursor.execute(f"UPDATE employee SET {update_clause} WHERE id = %s", values)
        conn.commit()
        conn.close()

        if changes:
            flash("อัปเดตข้อมูลเรียบร้อยแล้ว:<br>" + "<br>".join(changes), "info")
        else:
            flash("ไม่มีการเปลี่ยนแปลงข้อมูล", "warning")

        next_url = request.form.get("next_url") or url_for('index')
        return redirect(next_url)

    # กรณี GET ดึงประวัติการแก้ไข
    cursor.execute("SELECT * FROM employee_history WHERE employee_id = %s ORDER BY updated_at DESC", (emp_id,))
    history = cursor.fetchall()

    # เปรียบเทียบความเปลี่ยนแปลงในประวัติ (แสดงความแตกต่างระหว่าง record)
    for i in range(len(history) - 1):
        old = history[i + 1]
        new = history[i]
        changes = []
        for field in field_labels.keys():
            if field in ["birth", "start_date", "exit_date"]:
                old_val = format_thai_date(old[field]) if old[field] else ""
                new_val = format_thai_date(new[field]) if new[field] else ""
            elif field == "department_id":
                old_val = dept_dict.get(str(old[field]), "ไม่ระบุ")
                new_val = dept_dict.get(str(new[field]), "ไม่ระบุ")
            else:
                old_val = old[field]
                new_val = new[field]

            if old_val != new_val:
                label = field_labels.get(field, field)
                changes.append(f"{label}: '{old_val}' → '{new_val}'")

        history[i]["changes"] = "<br>".join(changes)

    if history:
        history[-1]["changes"] = "—"

    # แปลงวันที่ของพนักงานและประวัติให้เป็นฟอร์แมตไทย
    employee["birth"] = format_thai_date(employee["birth"])
    employee["start_date"] = format_thai_date(employee["start_date"])
    employee["exit_date"] = format_thai_date(employee["exit_date"])
    for h in history:
        h["birth"] = format_thai_date(h["birth"])
        h["start_date"] = format_thai_date(h["start_date"])
        h["exit_date"] = format_thai_date(h["exit_date"])

    conn.close()
    return render_template('update.html', employee=employee, history=history, departments=departments)

@app.route('/departments')
def departments():
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)
    cursor.execute("SELECT * FROM departments")
    departments = cursor.fetchall()
    conn.close()
    return render_template('departments.html', departments=departments)

@app.route('/add_department', methods=['POST'])
def add_department():
    name = request.form.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'กรุณาระบุชื่อแผนก'})

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO departments (name) VALUES (%s)", (name,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'เพิ่มแผนกเรียบร้อยแล้ว'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'เกิดข้อผิดพลาด: {e}'})

@app.route('/edit_department', methods=['POST'])
def edit_department():
    dept_id = request.form.get('id')
    new_name = request.form.get('name')
    if not dept_id or not new_name:
        return jsonify({'success': False, 'message': 'ข้อมูลไม่ครบถ้วน'})

    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE departments SET name = %s WHERE id = %s", (new_name, dept_id))
        conn.commit()
        message = "แก้ไขชื่อแผนกสำเร็จ"
        success = True
    except Exception as e:
        conn.rollback()
        message = f"เกิดข้อผิดพลาด: {str(e)}"
        success = False
    finally:
        cursor.close()
        conn.close()

    return jsonify({'success': success, 'message': message})

@app.route('/delete_department/<int:dept_id>', methods=['POST'])
def delete_department(dept_id):
    conn = get_connection()
    cursor = conn.cursor()

    try:
        # กำหนด id ของแผนก "ไม่มีแผนก"
        no_dept_id = 1

        # อัปเดต employee ให้ department_id เป็น no_dept_id แทนแผนกที่จะลบ
        cursor.execute("UPDATE employee SET department_id = %s WHERE department_id = %s", (no_dept_id, dept_id))

        # ถ้ามีตารางอื่นที่ FK ไปยัง departments ก็ทำแบบเดียวกัน

        # ลบแผนก
        cursor.execute("DELETE FROM departments WHERE id = %s", (dept_id,))

        conn.commit()
        return jsonify({"success": True, "message": "ลบแผนกสำเร็จ และอัปเดตข้อมูลที่เกี่ยวข้องเรียบร้อย"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": f"เกิดข้อผิดพลาด: {str(e)}"})
    finally:
        cursor.close()
        conn.close()

# ==================== fulltable ====================
@app.route('/full_table')
def full_table():
    conn = get_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # 1. รับค่าค้นหาจาก query string
    name = request.args.get('name', '').strip()
    gender = request.args.get('gender', '')
    department = request.args.get('department', '')

    # 2. สร้าง query + เงื่อนไข filter
    sql = """
        SELECT e.*, d.name AS department_name
        FROM employee e
        LEFT JOIN departments d ON e.department_id = d.id
        WHERE 1=1
    """
    filters = []
    values = []

    if name:
        filters.append("(e.name LIKE %s OR e.sur_name LIKE %s)")
        values.extend([f"%{name}%", f"%{name}%"])

    if gender:
        filters.append("e.gender = %s")
        values.append(gender)

    if department:
        filters.append("d.name = %s")
        values.append(department)

    if filters:
        sql += " AND " + " AND ".join(filters)

    sql += " ORDER BY e.id"

    cursor.execute(sql, values)
    employee = cursor.fetchall()

    # ดึงชื่อแผนกทั้งหมดสำหรับ dropdown
    cursor.execute("SELECT name FROM departments ORDER BY name")
    departments = [row["name"] for row in cursor.fetchall()]
    conn.close()

    # prepare และส่งไป template
    employee = prepare_employee_data(employee)

    return render_template("full_table.html",
                           employee=employee,
                           departments=departments,
                           search_params={"name": name, "gender": gender, "department": department})

@app.route('/export')
def export():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT e.id, e.pre, e.name, e.sur_name, e.type, d.name AS department,
               e.id_card, e.gender, e.nationality, e.senior, e.birth, e.start_date, e.phone, 
               e.email, e.address, e.bank_account, e.exit_date, e.exit_reason,
               e.main_skill, e.sub_skill1, e.sub_skill2
        FROM employee e
        LEFT JOIN departments d ON e.department_id = d.id
    """)
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        flash("ไม่มีข้อมูลให้ Export", "warning")
        return redirect(url_for('index'))

    # สร้าง DataFrame
    col_names = [
        "ID", "คำนำหน้า", "ชื่อ", "นามสกุล", "ประเภท", "แผนก", "เลขบัตรประชาชน",
        "เพศ", "สัญชาติ", "วุฒิ", "วันเกิด", "วันเริ่มงาน", "เบอร์โทร", "Email", "ที่อยู่",
        "เลขบัญชี", "วันออก", "หมายเหตุ",
        "ความสามารถหลัก", "ความสามารถอื่น 1", "ความสามารถอื่น 2"
    ]
    df = pd.DataFrame(rows, columns=col_names)

    # แปลงวันเป็น datetime
    df["วันเกิด"] = pd.to_datetime(df["วันเกิด"], errors='coerce')
    df["วันเริ่มงาน"] = pd.to_datetime(df["วันเริ่มงาน"], errors='coerce')
    df["วันออก"] = pd.to_datetime(df["วันออก"], errors='coerce')

    # เพิ่มคอลัมน์วันแบบข้อความ dd/mm/พ.ศ.
    df["วันเริ่มงาน (ข้อความ)"] = df["วันเริ่มงาน"].apply(
        lambda d: d.strftime("%d/%m/") + str(d.year + 543) if pd.notnull(d) else ""
    )
    df["วันออก (ข้อความ)"] = df["วันออก"].apply(
        lambda d: d.strftime("%d/%m/") + str(d.year + 543) if pd.notnull(d) else ""
    )

    # เพิ่มคอลัมน์ อายุงาน (ปีทศนิยม)
    def calculate_years_decimal(start, end=None):
        end = end or date.today()
        delta_days = (end - start).days
        return round(delta_days / 365.25, 2)

    years_decimal = []
    for idx, row in df.iterrows():
        sdate = row["วันเริ่มงาน"]
        edate = row["วันออก"]
        if pd.notnull(sdate):
            years_decimal.append(calculate_years_decimal(sdate.date(), edate.date() if pd.notnull(edate) else None))
        else:
            years_decimal.append("")
    df["อายุงาน (ปีทศนิยม)"] = years_decimal

    # เตรียมไฟล์
    filename = f"employee_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("static", filename)
    os.makedirs("static", exist_ok=True)

    # เตรียม workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "พนักงาน"

    # เพิ่มหัวตาราง
    headers = list(df.columns) + ["อายุงาน (สูตร)"]
    ws.append(headers)

    # เขียนข้อมูลลง Excel
    for idx, row in df.iterrows():
        values = list(row.values) + [None]  # ช่องสุดท้ายสำหรับสูตร
        ws.append(values)

    # ใส่ฟอร์แมตวันที่ และสูตร DATEDIF แบบใช้ข้อความ
    for i in range(2, len(df) + 2):  # แถว 2 เป็นต้นไป
        for col_letter in ['K', 'L', 'Q']:  # วันเกิด, เริ่มงาน, ออก
            cell = ws[f"{col_letter}{i}"]
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = '[$-th-TH]dd/mm/yyyy;@'

        # ใส่สูตร DATEDIF (โดยอิงจากข้อความ)
        start_text_cell = f"Z{i}"  # วันเริ่มงาน (ข้อความ)
        end_text_cell = f"AA{i}"  # วันออก (ข้อความ)
        formula = (
            f'=IF({start_text_cell}="", "", '
            f'DATEDIF(DATEVALUE({start_text_cell}), IF({end_text_cell}="", TODAY(), DATEVALUE({end_text_cell})), "Y") & "ปี" & '
            f'DATEDIF(DATEVALUE({start_text_cell}), IF({end_text_cell}="", TODAY(), DATEVALUE({end_text_cell})), "YM") & "เดือน" & '
            f'DATEDIF(DATEVALUE({start_text_cell}), IF({end_text_cell}="", TODAY(), DATEVALUE({end_text_cell})), "MD") & "วัน")'
        )
        col_formula = len(headers)  # คอลัมน์ "อายุงาน (สูตร)"
        ws.cell(row=i, column=col_formula, value=formula)

    # ปรับฟอนต์หัวตาราง
    for cell in ws[1]:
        cell.font = Font(name='TH Sarabun New', bold=True)

    # ตั้งค่าความกว้างเฉพาะบางคอลัมน์
    fixed_widths = {
        'A': 6,  # ID
        'B': 8,  # คำนำหน้า
        'I': 10,  # สัญชาติ
        'X': 22,  # อายุงาน (สูตร) - จำกัดให้พอดี
    }

    # ปรับความกว้าง
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter in fixed_widths:
            ws.column_dimensions[col_letter].width = fixed_widths[col_letter]
        else:
            # อัตโนมัติแบบจำกัดไม่เกิน 40
            max_len = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # บันทึก
    wb.save(filepath)

    flash(f"Export สำเร็จ: <a href='/{filepath}' target='_blank'>{filename}</a>", "success")
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000, debug=True)
